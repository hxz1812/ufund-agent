from dotenv import load_dotenv
import sys
import argparse
from llama_cpp import Llama
from pathlib import Path

import json
import time
import io
import os
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import (
    TextLoader,
    PyPDFLoader,
    Docx2txtLoader,
    CSVLoader,
)
from langchain_community.vectorstores import Chroma
from langchain_openai import OpenAIEmbeddings, ChatOpenAI
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import tiktoken
import numpy as np
from langchain_core.load import dumps, loads

from typing import Any, List, Optional, Callable
import shutil

from openpyxl import load_workbook
from langchain_core.documents import Document

import hashlib
import math

from datetime import datetime, timezone


model_path = str(Path.home() / "Desktop" / "agents-from-scratch" / "models" / "llama-3-8b-instruct.gguf")

stop = ["<|eot_id|>","User:"]


SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

FOLDER_ID = "1OqyC8z5ECyFALzjpQLE-vaNyNI5T2ey0"

BASE_DIR = Path(__file__).resolve().parent

ENV_PATH = BASE_DIR / ".env"
OUTPUT_DIR = BASE_DIR / "exported_docs_for_rag"
CHROMA_DIR = BASE_DIR / "chroma_db"

CHROMA_COLLECTION = "ufund_docs"
EMBEDDING_MODEL = "text-embedding-3-small"

REINDEX_STATE_PATH = (
    Path(__file__).resolve().parent
    / "reindex_state.json"
)

INDEX_BATCH_SIZE = 1000
CHUNK_SIZE = 300
CHUNK_OVERLAP = 50


GOOGLE_EXPORT_TYPES = {
    "application/vnd.google-apps.document": {
        "mime_type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "extension": ".docx",
    },
    "application/vnd.google-apps.spreadsheet": {
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "extension": ".xlsx",
    },
    "application/vnd.google-apps.presentation": {
        "mime_type": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "extension": ".pptx",
    },
    "application/vnd.google-apps.drawing": {
        "mime_type": "image/png",
        "extension": ".png",
    },
}

load_dotenv(dotenv_path=ENV_PATH)

RAG_ANSWER_SCHEMA = """
{
  "answer": "string",
  "evidence_files": ["string"],
  "confidence": "high | medium | low",
  "missing_information": ["string"]
}
"""

TOOL_CHOICES = [
    "calculator",
    "search_docs",
    "summarize_docs",
    "answer_from_docs",
    "none",
]

DROPBOX_DIR_TEXT = os.getenv("DROPBOX_DIR")

DROPBOX_DIR = (
    Path(DROPBOX_DIR_TEXT).expanduser()
    if DROPBOX_DIR_TEXT
    else None
)

ToolHandler = Callable[[dict, Any], Any]

DEFAULT_SEARCH_K = 4
SUMMARY_PAGE_SIZE = 500
TRACE_DIR = BASE_DIR / "local_traces"

def main() -> None:
    args = parse_args()
    
    user_input=args.prompt
    max_tokens=args.max_tokens
    temperature=args.temperature
    n_ctx=args.n_ctx
    max_steps=args.max_steps
    export_drive=args.export_drive
    llm_provider=args.llm_provider
    openai_model=args.openai_model
    reindex=args.reindex
    trace=args.trace

    system_prompt=("You are an assistant who truthfully and thoughtfully answers questions the members of UFund Investment LLC have.")

    if export_drive and not reindex:
        raise ValueError(
            "--export-drive must be used together with --reindex."
        )

    if not os.getenv("OPENAI_API_KEY"):
        raise RuntimeError(
            "Missing OPENAI_API_KEY. "
            "It is required for OpenAIEmbeddings."
        )

    embeddings = OpenAIEmbeddings(
        model=EMBEDDING_MODEL
    )

    if not reindex and not CHROMA_DIR.exists():
        raise RuntimeError(
            "No Chroma database exists. "
            "Run with --reindex first."
        )

    if reindex:
        vectorstore = reindex_documents(
            embeddings=embeddings,
            export_drive=export_drive,
        )
        if not user_input:
            return
    else:
        vectorstore = Chroma(
            collection_name=CHROMA_COLLECTION,
            persist_directory=str(CHROMA_DIR),
            embedding_function=embeddings,
        )

    if not user_input:
        raise ValueError(
            "--prompt is required unless you are using --reindex."
        )

    llm_config = build_llm(
        llm_provider=llm_provider,
        model_path=model_path,
        n_ctx=n_ctx,
        temperature=temperature,
        openai_model=openai_model,
    )

    agent_result = run_loop(
        llm_config=llm_config,
        system_prompt=system_prompt,
        user_input=user_input,
        vectorstore=vectorstore,
        max_steps=max_steps,
        max_tokens=max_tokens,
        temperature=temperature,
        trace=trace
    )
    print(json.dumps(agent_result, indent=2, ensure_ascii=False))

'''
    #multi query
    template = """You are an AI language modle assistant. Your task is to generate
five different versions of the given user question to retrieve relevant documents
from a vector database. By generating multiple perspectives on the user question,
your goal is to help the user overcome some of the limitations of the distance-based
similarity search. Provide these alternative questions separated by newlines.
Original question: {question}"""
    prompt_perspectives=ChatPromptTemplate.from_template(template)
    generate_queries=(
        prompt_perspectives
        | ChatOpenAI(temperature=0)
        | StrOutputParser()
        | (lambda x: x.split("\n"))
        )
    retrieval_chain = generate_queries | retriever.map() | get_unique_union
    docs = retrieval_chain.invoke({"question":user_input})
    #print(len(docs))'''

##############
# PARSE ARGS #
##############
def parse_args():
    parser=argparse.ArgumentParser()
    parser.add_argument("--prompt", help="Question to ask the agent.")
    parser.add_argument("--max-tokens", type=int, default=512)
    parser.add_argument("--temperature", type=float, default=0)
    parser.add_argument("--n-ctx",type=int,default=2048)
    parser.add_argument("--max-steps", type=int, default=5)
    parser.add_argument("--export-drive", action="store_true")
    parser.add_argument("--llm-provider", choices=["local","openai"],default="local",)
    parser.add_argument("--openai-model", default="gpt-3.5-turbo",)
    parser.add_argument("--reindex", action="store_true")
    parser.add_argument("--trace", action="store_true", help="Save agent messages, actions, and tool outputs locally.")
    return parser.parse_args()

###############
# GETTING LLM #
###############
def build_llm(llm_provider,model_path,n_ctx,temperature,openai_model):
    if llm_provider=="local":
        return {
            "provider": "local",
            "client": Llama(
                model_path=str(model_path),
                temperature=temperature,
                n_ctx=n_ctx,
                verbose=False,
                ),
            }
    if llm_provider=="openai":
        if not os.getenv("OPENAI_API_KEY"):
            raise RuntimeError("Missing OPENAI_API_KEY")
        return {
            "provider": "openai",
            "client": ChatOpenAI(
                model_name=openai_model,
                temperature=temperature,
                ),
            }
    raise ValueError(f"Unsupported LLM provider: {llm_provider}")

def call_llm_text(llm_config, prompt, max_tokens=512, temperature=0, stop=None,):
    provider=llm_config["provider"]
    client = llm_config["client"]
    if provider == "local":
        response= client(prompt=prompt,max_tokens=max_tokens,temperature=temperature,stop=stop or ["<|eot_id|>", "User:"])
        return response["choices"][0]["text"].strip()
    if provider=="openai":
        response=client.invoke(prompt)
        return response.content.strip()
    raise ValueError(f"Unsupported LLM provider: {provider}")
        
    
##########################
#    GDRIVE FUNCTIONS    #
##########################
def get_drive_service():
    creds=None
    if os.path.exists("token.json"):
        creds=Credentials.from_authorized_user_file("token.json", SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow=InstalledAppFlow.from_client_secrets_file(
                "credentials.json",
                SCOPES,
                )
            creds=flow.run_local_server(port=0)
        with open("token.json","w") as token:
            token.write(creds.to_json())
    return build("drive","v3",credentials=creds)

def safe_filename(name):
    bad_chars=['/','\\',':','*','?','"','<','>','|']
    for char in bad_chars:
        name=name.replace(char,"_")
    return name.strip()

def list_files_in_folder(service,folder_id):
    files=[]
    page_token=None
    while True:
        response=(service.files().list(
            q=f"'{folder_id}' in parents and trashed = false",
            spaces="drive",
            fields="nextPageToken, files(id, name, mimeType)",
            pageToken=page_token,).execute()
                  )
        files.extend(response.get("files",[]))
        page_token=response.get("nextPageToken")
        if not page_token:
            break
    return files

def export_google_file(service,file_id,file_name,export_info,output_path):
    request=service.files().export_media(
        fileId=file_id,
        mimeType=export_info["mime_type"]
        )
    with io.FileIO(output_path,"wb") as file_handle:
        downloader=MediaIoBaseDownload(file_handle,request)
        done=False
        while not done:
            status,done=downloader.next_chunk()

def download_regular_file(service,file_id,output_path):
    request=service.files().get_media(fileId=file_id)
    with io.FileIO(output_path,"wb") as file_handle:
        downloader=MediaIoBaseDownload(file_handle,request)
        done=False
        while not done:
            status,done=downloader.next_chunk()

def export_folder(service,folder_id,output_dir,relative_path=""):
    output_dir=Path(output_dir)
    current_output_dir=output_dir/relative_path
    current_output_dir.mkdir(parents=True,exist_ok=True)
    files=list_files_in_folder(service,folder_id)
    exported=[]
    downloaded=[]
    skipped=[]
    for file in files:
        file_id=file["id"]
        file_name=file["name"]
        mime_type=file["mimeType"]
        clean_name=safe_filename(file_name)
        if mime_type=="application/vnd.google-apps.folder":
            print(f"Entering folder: {file_name}")
            child_exported,child_downloaded,child_skipped=export_folder(
                service=service,
                folder_id=file_id,
                output_dir=output_dir,
                relative_path=str(Path(relative_path)/clean_name),
                )
            exported.extend(child_exported)
            downloaded.extend(child_downloaded)
            skipped.extend(child_skipped)
            continue
        try:
            if mime_type in GOOGLE_EXPORT_TYPES:
                export_info= GOOGLE_EXPORT_TYPES[mime_type]
                output_path=current_output_dir/f"{clean_name}{export_info['extension']}"
                export_google_file(
                    service=service,
                    file_id=file_id,
                    file_name=file_name,
                    export_info=export_info,
                    output_path=output_path,
                    )
                exported.append(str(output_path))
                print(f"Exported: {output_path}")
            elif mime_type.startswith("application/vnd.google-apps."):
                skipped.append(f"{file_name} | unsupported Google Workspace type: {mime_type}")
                print(f"Skipped: {file_name}")
            else:
                output_path=current_output_dir/clean_name
                download_regular_file(
                    service=service,
                    file_id=file_id,
                    output_path=output_path,
                    )
                downloaded.append(str(output_path))
                print(f"Downloaded: {output_path}")
        except Exception as e:
            skipped.append(f"{file_name} | ERROR: {e}")
            print(f"Error: {file_name} | {e}")
    return exported,downloaded,skipped

############
# INDEXING #
############
def load_xlsx_documents(path: Path, rows_per_document: int = 50,) -> list[Document]:
    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )
    documents = []

    try:
        for worksheet in workbook.worksheets:
            batch_lines = []
            batch_start_row = None
            for row_number, row in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                values = [
                    "" if value is None else str(value)
                    for value in row
                ]
                
                if not any(value.strip() for value in values):
                    continue
                if batch_start_row is None:
                    batch_start_row = row_number
                batch_lines.append("\t".join(values))
                if len(batch_lines) >= rows_per_document:
                    documents.append(
                        Document(
                            page_content=(
                                f"Worksheet: {worksheet.title}\n"
                                f"Rows: {batch_start_row}-{row_number}\n\n"
                                + "\n".join(batch_lines)
                            ),
                            metadata={
                                "sheet_name": worksheet.title,
                                "row_start": batch_start_row,
                                "row_end": row_number,
                            },
                        )
                    )
                    batch_lines = []
                    batch_start_row = None
            if batch_lines:
                documents.append(
                    Document(
                        page_content=(
                            f"Worksheet: {worksheet.title}\n"
                            f"Rows: {batch_start_row}-{worksheet.max_row}\n\n"
                            + "\n".join(batch_lines)
                        ),
                        metadata={
                            "sheet_name": worksheet.title,
                            "row_start": batch_start_row,
                            "row_end": worksheet.max_row,
                        },
                    )
                )
    finally:
        workbook.close()
    return documents

def load_documents(DOCS_PATH,printing=False,source_system="local"):
    docs=[]
    skipped_files=[]
    root_path = Path(DOCS_PATH).expanduser().resolve()
    for path in sorted(root_path.rglob("*")):
        if path.is_dir() or path.name.startswith('.'):
            continue
        suffix = path.suffix.lower()
        try:
            if suffix == ".xlsx":
                loaded_docs = load_xlsx_documents(
                    path=path,
                    rows_per_document=15,
                )
            else:
                if suffix in ['.txt', '.md']:
                    loader=TextLoader(str(path),encoding='utf-8')
                elif suffix == '.pdf':
                    loader = PyPDFLoader(str(path))
                elif suffix=='.docx':
                    loader=Docx2txtLoader(str(path))
                elif suffix=='.csv':
                    loader=CSVLoader(str(path))
                else:
                    skipped_files.append(
                        f"{path} | unsupported file type"
                    )
                    continue

                loaded_docs = loader.load()
            for doc in loaded_docs:
                resolved_path = path.resolve()
                original_content = doc.page_content.strip()
                doc.metadata["source"] = str(resolved_path)
                doc.metadata["source_system"] = source_system
                doc.metadata["relative_path"] = str(
                    resolved_path.relative_to(root_path)
                )
                doc.metadata["file_name"] = path.name
                doc.metadata["file_type"] = suffix
                doc.page_content = (
                    f"File name: {path.name}\n\n"
                    f"{original_content}"
                )
            docs.extend(loaded_docs)
        except Exception as e:
            skipped_files.append(f"{path} | ERROR: {e}")
    if printing:
        print("Loaded docs:", len(docs))
        print("Skipped files:", len(skipped_files))
        if docs:
            print("First doc metadata:")
            print(docs[0].metadata)
            print("Loaded file names:")
            for name in sorted(set(doc.metadata.get("file_name") for doc in docs)):
                print(name)
        if skipped_files:
            print("Skipped:")
            for item in skipped_files[:20]:
                print(item)
    return docs, skipped_files

#checkpoint function
def load_reindex_state() -> dict | None:
    if not REINDEX_STATE_PATH.exists():
        return None
    try:
        return json.loads(
            REINDEX_STATE_PATH.read_text(
                encoding="utf-8"
            )
        )
    except (json.JSONDecodeError, OSError):
        return None

#checkpoint function
def save_reindex_state(state: dict) -> None:
    temporary_path = REINDEX_STATE_PATH.with_suffix(
        ".json.tmp"
    )
    temporary_path.write_text(
        json.dumps(
            state,
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    temporary_path.replace(REINDEX_STATE_PATH)

def create_chunk_ids(splits,) -> list[str]:
    source_chunk_numbers = {}
    chunk_ids = []
    for chunk in splits:
        source_system = chunk.metadata.get(
            "source_system",
            "unknown",
        )
        source = chunk.metadata.get(
            "source",
            "unknown",
        )
        source_key = f"{source_system}|{source}"
        chunk_number = source_chunk_numbers.get(
            source_key,
            0,
        )
        source_chunk_numbers[source_key] = (
            chunk_number + 1
        )
        content_hash = hashlib.sha256(
            chunk.page_content.encode("utf-8")
        ).hexdigest()
        raw_id = (
            f"{source_key}|"
            f"{chunk_number}|"
            f"{content_hash}"
        )
        chunk_id = hashlib.sha256(
            raw_id.encode("utf-8")
        ).hexdigest()
        chunk.metadata["chunk_number"] = chunk_number
        chunk.metadata["content_hash"] = content_hash
        chunk_ids.append(chunk_id)
    return chunk_ids

#check doc / chunk config changes
def create_index_signature(chunk_ids: list[str]) -> str:
    digest = hashlib.sha256()
    digest.update(
        EMBEDDING_MODEL.encode("utf-8")
    )
    digest.update(
        str(CHUNK_SIZE).encode("utf-8")
    )
    digest.update(
        str(CHUNK_OVERLAP).encode("utf-8")
    )
    for chunk_id in chunk_ids:
        digest.update(
            chunk_id.encode("utf-8")
        )

    return digest.hexdigest()

def reindex_documents(embeddings, export_drive: bool = False,):
    if export_drive:
        service = get_drive_service()

        exported, downloaded, skipped = export_folder(
            service=service,
            folder_id=FOLDER_ID,
            output_dir=OUTPUT_DIR,
        )

        print("Exported:", len(exported))
        print("Downloaded:", len(downloaded))
        print("Skipped during export:", len(skipped))

    docs = []
    skipped_files = []

    document_sources = [
        ("google_drive", OUTPUT_DIR),
    ]

    if DROPBOX_DIR is not None:
        if not DROPBOX_DIR.exists():
            raise RuntimeError(
                f"Dropbox directory does not exist: "
                f"{DROPBOX_DIR}"
            )

        if not DROPBOX_DIR.is_dir():
            raise RuntimeError(
                f"DROPBOX_DIR is not a folder: "
                f"{DROPBOX_DIR}"
            )

        document_sources.append(
            ("dropbox", DROPBOX_DIR)
        )

    for source_system, source_path in document_sources:
        source_path = Path(source_path).expanduser()

        if not source_path.exists():
            print(
                f"Skipping missing document source: "
                f"{source_system} — {source_path}"
            )
            continue

        print(
            f"\nLoading {source_system} documents from:"
            f"\n{source_path.resolve()}"
        )

        source_docs, source_skipped = load_documents(
            DOCS_PATH=source_path,
            printing=True,
            source_system=source_system,
        )

        docs.extend(source_docs)
        skipped_files.extend(source_skipped)

    if not docs:
        raise RuntimeError(
            "No supported documents were found in "
            "Google Drive exports or Dropbox."
        )

    text_splitter = (
        RecursiveCharacterTextSplitter
        .from_tiktoken_encoder(
            chunk_size=CHUNK_SIZE,
            chunk_overlap=CHUNK_OVERLAP,
        )
    )
    splits = []

    for doc in docs:
        if doc.metadata.get("file_type") == ".xlsx":
            splits.append(doc)
        else:
            splits.extend(
                text_splitter.split_documents([doc])
            )

    if not splits:
        raise RuntimeError(
            "Documents were loaded, but no chunks were created."
        )

    chunk_ids = create_chunk_ids(splits)

    index_signature = create_index_signature(
        chunk_ids
    )

    state = load_reindex_state()

    can_resume = (
        state is not None
        and state.get("signature") == index_signature
        and CHROMA_DIR.exists()
    )

    if can_resume:
        start_batch = int(
            state.get("next_batch", 0)
        )
        print(
            "Resuming interrupted reindex from "
            f"batch {start_batch + 1}."
        )

    else:
        if CHROMA_DIR.exists():
            shutil.rmtree(CHROMA_DIR)

        start_batch = 0

        state = {
            "signature": index_signature,
            "next_batch": 0,
            "total_batches": math.ceil(
                len(splits) / INDEX_BATCH_SIZE
            ),
            "complete": False,
        }
        save_reindex_state(state)
        print("Starting a new reindex job.")

    vectorstore = Chroma(
        collection_name=CHROMA_COLLECTION,
        persist_directory=str(CHROMA_DIR),
        embedding_function=embeddings,
    )


    total_batches = math.ceil(
        len(splits) / INDEX_BATCH_SIZE
    )

    if state.get("complete"):
        print(
            "This exact document index is already complete."
        )
        return vectorstore

    for batch_number in range(
        start_batch,
        total_batches,
    ):
        start = batch_number * INDEX_BATCH_SIZE
        end = min(
            start + INDEX_BATCH_SIZE,
            len(splits),
        )
        batch_documents = splits[start:end]
        batch_ids = chunk_ids[start:end]
        print(
            f"Indexing batch {batch_number + 1} "
            f"of {total_batches}: "
            f"chunks {start + 1}-{end}"
        )
        vectorstore.add_documents(
            documents=batch_documents,
            ids=batch_ids,
        )
        state["next_batch"] = batch_number + 1
        save_reindex_state(state)
        
    state["complete"] = True
    state["next_batch"] = total_batches
    save_reindex_state(state)

    print(
        f"Finished indexing {len(splits)} chunks into "
        f"{CHROMA_DIR.resolve()}."
    )
    print(
        f"Skipped {len(skipped_files)} unsupported "
        f"or failed files."
    )

    return vectorstore

def num_tokens_from_string(string: str, encoding_name: str) -> int:
    encoding=tiktoken.get_encoding(encoding_name)
    num_tokens=len(encoding.encode(string))
    return num_tokens

def cosine_similarity(vec1,vec2):
    dot_product=np.dot(vec1,vec2)
    norm_vec1=np.linalg.norm(vec1)
    norm_vec2=np.linalg.norm(vec2)
    return dot_product/(norm_vec1*norm_vec2)

def format_docs(retrieved_docs):
    return "\n\n".join(doc.page_content for doc in retrieved_docs)

##################################
# ACTUALLY RUNNING RAG + TRACING #
##################################
def run_rag_with_local_trace(question):
    retrieved_docs=retriever.invoke(question)
    context= "\n\n".join(doc.page_content for doc in retrieved_docs)
    answer = answer_chain.invoke({
        "context": context,
        "question": question,
        })
    trace = {
        "question": question,
        "answer": answer,
        "retrieved_docs": [
            {
                "file_name": doc.metadata.get("file_name"),
                "source": doc.metadata.get("source"),
                "file_type": doc.metadata.get("file_type"),
                "preview": doc.page_content[:1000],
                }
            for doc in retrieved_docs
            ],
        "context_preview": context[:3000],
        }
    trace_path = TRACE_DIR / f"rag_trace{int(time.time())}.json"
    with open(trace_path, "w", encoding = "utf-8") as f:
        json.dump(trace, f, indent=2, ensure_ascii=False)
    print("Saved local trace to:", trace_path)
    return answer, retrieved_docs

def answer_with_local_llm(llm, user_input, retrieved_docs, system_prompt, max_tokens=512, temperature=0):
    context = "\n\n".join(doc.page_content for doc in retrieved_docs)
    prompt = f"""{system_prompt}

CRITICAL INSTRUCTIONS:
1. Everything in your response will have evidence to support it from given documents and text.
2. If you cannot find evidence or facts, you will respond with "I don't know."
3. Keep the answer as brief as possible. Do not make things up, do not provide extra explanations.

Question: {user_input}

Context:
{context}

Answer:"""

    response = llm(prompt,max_tokens=max_tokens,temperature=temperature,)
    return response["choices"][0]["text"].strip()

#union of retrieved docs
def get_unique_union(documents: list[list]):
    flattened_docs=[dumps(doc) for sublist in documents for doc in sublist]
    unique_docs=list(set(flattened_docs))
    return [loads(doc) for doc in unique_docs]

#####################
# STRUCTURED OUTPUT #
#####################
def extract_json_from_text(text: str) -> dict | None:
    text = text.strip()
    decoder=json.JSONDecoder()

    for i, char in enumerate(text):
        if char != "{":
            continue

        try:
            parsed, _ = decoder.raw_decode(text[i:])
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            continue

    return None


def generate_structured(
    llm_config,
    user_input: str,
    schema: str,
    system_prompt: str,
    max_tokens: int = 512,
    temperature: float = 0.0,
    stop= None,
) -> dict | None:

    prompt = f"""{system_prompt}

CRITICAL INSTRUCTIONS:
1. Respond with ONLY valid JSON.
2. No explanations, no markdown, no extra text before or after the JSON.
3. Start your response with {{ and end with }}.
4. Use only the retrieved context provided by the user request.
5. If the answer is not supported by the context, set "answer" to "I don't know."

Schema you must follow:
{schema}

User request:
{user_input}

Response (JSON only):"""

    for attempt in range(3):
        response = call_llm_text(llm_config=llm_config,prompt=prompt,max_tokens=max_tokens,temperature=temperature,stop=stop)
        parsed = extract_json_from_text(response)
        if parsed is not None:
            return parsed

    return None


def answer_rag_structured(
    llm_config,
    user_input: str,
    retrieved_docs,
    system_prompt: str,
    schema: str = RAG_ANSWER_SCHEMA,
    max_tokens: int = 512,
    temperature: float = 0.0,):
    context = "\n\n".join(
        f"File name: {doc.metadata.get('file_name')}\n"
        f"Source: {doc.metadata.get('source')}\n"
        f"Content:\n{doc.page_content}"
        for doc in retrieved_docs
    )

    structured_user_input = f"""Question:
{user_input}

Retrieved context:
{context}

Instructions:
- Answer using only the retrieved context.
- Do not use outside knowledge.
- If the retrieved context does not explicitly answer the question, set "answer" to "I don't know."
- Put source file names in "evidence_files".
- Use "confidence" as high, medium, or low.
- If context is insufficient, explain what is missing in "missing_information".
"""

    return generate_structured(
        llm_config=llm_config,
        user_input=structured_user_input,
        schema=schema,
        system_prompt=system_prompt,
        max_tokens=max_tokens,
        temperature=temperature,
    )

###################
# DECISION-MAKING #
###################
def decide(llm_config, system_prompt:str, user_input:str, choices: list[str],
           max_tokens:int=512,temperature:float=0.0, stop:List[str]=["<|eot_id|>","User:"])->str | None:
    options = "\n".join(f"- {choice}" for choice in choices)
    prompt = f"""{system_prompt}
You must choose ONE of the following options. Respond with ONLY valid JSON.

CRITICAL INSTRUCTIONS:
1. Response with ONLY valid JSON
2. No explanations, no markdown, no other text
3. Start your response with {{ and end with }}

Available choices:
{options}

Required JSON format:
{{"decision": "one_of_the_choices_above"}}

User request: {user_input}

Response (JSON only):"""

    for attempt in range(3):
        response = call_llm_text(llm_config=llm_config,prompt=prompt,
                                 max_tokens=max_tokens,temperature=temperature,
                                 stop=stop,)
        parsed = extract_json_from_text(response)
        if parsed and "decision" in parsed:
            decision = parsed["decision"]
            if decision in choices:
                return decision
    return None

#########
# TOOLS #
#########
def request_tool(llm_config, system_prompt:str, user_input:str, choices:list[str]=TOOL_CHOICES,max_tokens:int=512,temperature:float=0.0,
                 stop:list[str]=["<|eot_id|>", "<|end_of_text|>", "\nUser request:", "\nInput:"])->dict|None:
    options = "\n".join(f"- {choice}" for choice in choices)
    prompt=f"""{system_prompt}

You are a tool-calling assistant. When asked a math question, you must respond with ONLY valid JSON.

Available tools: {options}

Tool guidance:
- calculator: use for arithmetic/math calculations
- search_docs: use when the user asks to find information in documents
- summarize_docs: use when the user asks to summarize documents
- answer_from_docs: use when the user asks a question that should be answered from retrieved documents
- none: use when no tool is needed

CRITICAL INSTRUCTIONS:
1. Respond with ONLY valid JSON
2. No explanations, no markdown, no other text
3. Start your response with {{ and end with }}

Example output format:
{{"tool": "calculator", "arguments": {{"a": 42, "b": 7, "operation": "divide"}}}}

User request: {user_input}

Response (JSON only):"""
    for attempt in range(3):
        response=call_llm_text(llm_config=llm_config,prompt=prompt,
                               max_tokens=max_tokens,temperature=temperature,
                               stop=stop,)
        parsed=extract_json_from_text(response)
        if parsed and "tool" in parsed and "arguments" in parsed:
            return parsed
    return None

def execute_tool_call(tool_call:dict,vectorstore=None,)->Any:
    tool_name=tool_call.get("tool")
    arguments=tool_call.get("arguments",{}) or {}
    try:
        result=execute_tool(tool_name=tool_name,arguments=arguments,vectorstore=vectorstore,)
        return {"tool": tool_name, "arguments": arguments, "result": result, "success": True,}
    except Exception as e:
        return {"tool": tool_name, "arguments": arguments, "error": str(e), "success": False,}

def search_docs(vectorstore, query: str, k: int = DEFAULT_SEARCH_K):
    retrieved_docs = vectorstore.similarity_search(
        query,
        k=k,
    )
    results=[]
    for evidence_id, doc in enumerate(retrieved_docs, start=1):
        results.append({
            "evidence_id": evidence_id,
            "file_name": doc.metadata.get("file_name"),
            "source_system": doc.metadata.get(
                "source_system"
            ),
            "relative_path": doc.metadata.get(
                "relative_path"
            ),
            "source": doc.metadata.get("source"),
            "file_type": doc.metadata.get(
                "file_type"
            ),
            "sheet_name": doc.metadata.get(
                "sheet_name"
            ),
            "row_start": doc.metadata.get(
                "row_start"
            ),
            "row_end": doc.metadata.get(
                "row_end"
            ),
            "content": doc.page_content,
        })
    return results

def summarize_docs(vectorstore, limit:int=5):
    file_names = set()
    offset = 0
    while True:
        stored_page = vectorstore.get(
            include=["metadatas"],
            limit=page_size,
            offset=offset,
        )
        metadatas = (
            stored_page.get("metadatas")
            or []
        )
        if not metadatas:
            break
        for metadata in metadatas:
            if not metadata:
                continue
            file_names.add(
                metadata.get(
                    "file_name",
                    "unknown",
                )
            )
        if len(metadatas) < page_size:
            break
        offset += page_size
    sorted_file_names = sorted(file_names)
    return {
        "available_files": sorted_file_names[:limit],
        "file_count": len(sorted_file_names),
    }

def execute_search_docs(arguments: dict, vectorstore,):
    if vectorstore is None:
        raise ValueError(
            "vectorstore is required for search_docs"
        )
    return search_docs(
        vectorstore=vectorstore,
        query=arguments.get("query", ""),
        k=int(arguments.get("k", 4)),
    )


def execute_summarize_docs(arguments: dict, vectorstore,):
    if vectorstore is None:
        raise ValueError(
            "vectorstore is required for summarize_docs"
        )
    return summarize_docs(
        vectorstore=vectorstore,
        limit=int(arguments.get("limit", 5)),
    )


def execute_none(arguments: dict, vectorstore,):
    return "No tool used."

TOOL_REGISTRY = {
    "search_docs": {
        "description": (
            "Search company documents for information "
            "relevant to the user's question."
        ),
        "arguments": {
            "query": "string",
            "k": "integer, default 4",
        },
        "handler": execute_search_docs,
    },
    "summarize_docs": {
        "description": (
            "List or summarize the documents available "
            "in the vector database."
        ),
        "arguments": {
            "limit": "integer, default 5",
        },
        "handler": execute_summarize_docs,
    },
    "none": {
        "description": "Use when no tool is required.",
        "arguments": {},
        "handler": execute_none,
    },
}

def execute_tool(tool_name: str, arguments: dict, vectorstore=None):
    tool_spec = TOOL_REGISTRY.get(tool_name)
    if tool_spec is None:
        raise ValueError(
            f"Unknown tool: {tool_name}"
        )
    tool_handler = tool_spec.get("handler")
    if not callable(tool_handler):
        raise TypeError(
            f"Tool '{tool_name}' has no callable handler."
        )
    return tool_handler(
        arguments=arguments,
        vectorstore=vectorstore,
    )

def get_successful_tool_messages(messages: list[dict], tool_name: str | None = None,) -> list[dict]:
    return [message for message in messages if (
            message.get("role") == "tool"
            and message.get("success") is True
            and (tool_name is None or message.get("name") == tool_name)
            )]

def get_latest_successful_tool_payload(messages: list[dict], tool_name: str | None = None) -> dict | None:
    successful_messages = (get_successful_tool_messages(
            messages=messages,
            tool_name=tool_name,
        ))
    for message in reversed(successful_messages):
        content = message.get("content", {})
        if isinstance(content, dict):
            payload = content
        else:
            try:
                payload = json.loads(content)
            except (json.JSONDecodeError, TypeError):
                continue
        if isinstance(payload, dict):
            return payload
    return None

def get_latest_search_results(messages: list[dict]) -> list[dict]:
    payload = get_latest_successful_tool_payload(
        messages=messages,
        tool_name="search_docs",
    )
    if not payload:
        return []
    results = payload.get("result", [])
    return results if isinstance(results, list) else []

def get_documents_used(messages: list[dict], evidence_ids: list[int]) -> list[dict]:
    search_results = get_latest_search_results(messages)
    results_by_id = {
        result.get("evidence_id"): result
        for result in search_results
    }
    documents_used = []
    seen_sources = set()
    for evidence_id in evidence_ids:
        result = results_by_id.get(evidence_id)
        if result is None:
            continue
        source_key = (
            result.get("source_system"),
            result.get("relative_path"),
            result.get("sheet_name"),
            result.get("row_start"),
            result.get("row_end"),
        )
        if source_key in seen_sources:
            continue
        seen_sources.add(source_key)
        documents_used.append({
            "evidence_id": evidence_id,
            "file_name": result.get("file_name"),
            "source_system": result.get(
                "source_system"
            ),
            "relative_path": result.get(
                "relative_path"
            ),
            "file_type": result.get(
                "file_type"
            ),
            "sheet_name": result.get(
                "sheet_name"
            ),
            "row_start": result.get(
                "row_start"
            ),
            "row_end": result.get(
                "row_end"
            ),
        })
    return documents_used

########
# LOOP #
########
def render_messages(messages: list[dict])->str:
    rendered_messages = []

    for message in messages:
        role = message.get("role", "unknown")
        name = message.get("name")
        content = message.get("content", "")

        if not isinstance(content, str):
            content = json.dumps(
                content,
                ensure_ascii=False,
            )

        role_label = role.upper()

        if name:
            role_label += f" ({name})"

        rendered_messages.append(
            f"{role_label}:\n{content}"
        )

    return "\n\n".join(rendered_messages)

def render_tool_descriptions() -> str:
    sections = []

    for index, (tool_name, spec) in enumerate(
        TOOL_REGISTRY.items(),
        start=1,
    ):
        arguments_json = json.dumps(
            spec["arguments"],
            ensure_ascii=False,
        )

        sections.append(
            f"{index}. {tool_name}\n"
            f"   - {spec['description']}\n"
            f"   - Arguments: {arguments_json}"
        )

    return "\n\n".join(sections)

def agent_step(llm_config, messages:list[dict], system_prompt:str,
               max_tokens:int=128,temperature:float=0.0, max_parse_attempts:int=3,
                   stop=["<|eot_id|>", "<|end_of_text|>", "\n#", "\n\n"])-> dict | None:
    conversation = render_messages(messages)
    available_tools = render_tool_descriptions()
    successful_tool_messages = (get_successful_tool_messages(messages))
    latest_successful_tool_name = (
        successful_tool_messages[-1].get("name")
        if successful_tool_messages
        else None
    )
    search_results = get_latest_search_results(messages)
    available_evidence_ids = {
        result.get("evidence_id")
        for result in search_results
        if isinstance(
            result.get("evidence_id"),
            int,
        )
    }
    if successful_tool_messages:
        tool_status = (
            f"A successful result from "
            f"'{latest_successful_tool_name}' is already "
            "available. You MUST now return final_answer. "
            "Do not request another tool. Use only the "
            "existing tool evidence."
        )
    else:
        tool_status = (
            "No successful tool result is available. "
            "Use search_docs for questions about document "
            "content. Use summarize_docs only when the user "
            "asks which documents are available."
        )

    base_prompt = f"""{system_prompt}

You are an agent that may request one tool or return a final answer.

Available tools:
{available_tools}

CRITICAL INSTRUCTIONS:
1. Respond with ONLY one valid JSON object.
2. Do not include markdown or text outside the JSON.
3. Use only tools listed under Available tools.
4. Do not use outside knowledge for company-document questions.
5. Every factual claim must be supported by existing tool output.
6. Do not infer facts merely from filenames or folder paths.
7. If the evidence is insufficient, answer "I don't know."
8. Use search_docs for questions about document contents.
9. Use summarize_docs only to list or count available documents.
10. After one tool succeeds, return final_answer without calling another tool.
11. For search_docs evidence, cite evidence IDs as [1], [2], and so on.
12. Include only evidence_ids that directly support the answer.

Tool status:
{tool_status}

Required JSON formats:

For tool use:
{{"action": "tool_use", "tool": "tool_name", "arguments": {{}}, "reason": "why the tool is needed"}}

For final answer:
{{"action": "final_answer", "answer": "answer text with citations such as [1]", "evidence_ids": [1], "reason": "why the cited evidence supports the answer"}}

For an unsupported answer:
{{"action": "final_answer", "answer": "I don't know.", "evidence_ids": [], "reason": "the retrieved evidence is insufficient"}}

Conversation:
{conversation}

Response (JSON only):"""

    stop_sequences = stop or [
        "<|eot_id|>",
        "<|end_of_text|>",
    ]

    retry_feedback = ""
    for attempt in range(max_parse_attempts):
        current_prompt = (base_prompt + retry_feedback)
        response = call_llm_text(
            llm_config=llm_config,
            prompt=current_prompt,
            max_tokens=max_tokens,
            temperature=temperature,
            stop=stop_sequences,
        )
        parsed = extract_json_from_text(response)
        if not parsed:
            retry_feedback = f"""

RETRY FEEDBACK:
Your previous response was not valid JSON:

{response}

Return exactly one valid JSON object.
"""
            continue
        action = parsed.get("action")
        if action == "tool_use":
            tool_name = parsed.get("tool")
            arguments = parsed.get("arguments")
            if successful_tool_messages:
                retry_feedback = """

RETRY FEEDBACK:
A tool has already completed successfully.
Do not request another tool.
Return final_answer using the existing result.
"""
                continue

            if tool_name not in TOOL_REGISTRY:
                retry_feedback = f"""

RETRY FEEDBACK:
"{tool_name}" is not an available tool.
Choose one tool listed under Available tools.
"""
                continue

            if not isinstance(arguments, dict):
                retry_feedback = """

RETRY FEEDBACK:
The "arguments" field must be a JSON object.
"""
                continue

        elif action == "final_answer":
            answer = parsed.get("answer")
            evidence_ids = parsed.get("evidence_ids", [])

            if not isinstance(answer, str):
                retry_feedback = """

RETRY FEEDBACK:
A final_answer must contain an "answer" string.
"""
                continue

            if not successful_tool_messages:
                retry_feedback = """

RETRY FEEDBACK:
No successful tool result exists yet.
Request an appropriate tool before answering.
"""
                continue

            if not isinstance(evidence_ids, list):
                retry_feedback = """

RETRY FEEDBACK:
"evidence_ids" must be a JSON list.
"""
                continue

            if latest_successful_tool_name == "search_docs":
                invalid_evidence_ids = [
                    evidence_id
                    for evidence_id in evidence_ids
                    if (not isinstance(evidence_id,int)
                        or evidence_id
                        not in available_evidence_ids
                    )
                ]

                if invalid_evidence_ids:
                    retry_feedback = f"""

RETRY FEEDBACK:
Invalid evidence IDs were provided.
Available evidence IDs are:
{sorted(available_evidence_ids)}
"""
                    continue

                normalized_answer = (
                    answer.strip()
                    .lower()
                    .rstrip(".")
                )

                is_unknown_answer = (
                    normalized_answer.startswith(
                        "i don't know"
                    )
                    or normalized_answer.startswith(
                        "i do not know"
                    )
                )

                if (
                    not is_unknown_answer
                    and not evidence_ids
                ):
                    retry_feedback = """

RETRY FEEDBACK:
A factual answer based on search_docs must include
at least one valid evidence_id. Otherwise answer
"I don't know."
"""
                    continue

            else:
                parsed["evidence_ids"] = []

        else:
            retry_feedback = """

RETRY FEEDBACK:
The "action" field must be either:
- "tool_use"
- "final_answer"
"""
            continue

        parsed.setdefault(
            "reason",
            f"Taking action: {action}",
        )

        return parsed

    return None

def create_trace_path() -> Path:
    TRACE_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S_%fZ")
    return TRACE_DIR / (f"agent_trace_{timestamp}.jsonl")

def append_trace_event(trace_path: Path | None, event: dict) -> None:
    if trace_path is None:
        return
    trace_record = {
        "timestamp": datetime.now(
            timezone.utc
        ).isoformat(),
        **event,
    }
    with trace_path.open(
        "a",
        encoding="utf-8",
    ) as trace_file:
        trace_file.write(
            json.dumps(
                trace_record,
                ensure_ascii=False,
            )
            + "\n"
        )

def run_loop(llm_config, system_prompt:str, user_input:str, vectorstore=None,
             max_tokens:int=512,temperature:float=0.0,max_steps:int=5,trace:bool=False):
    messages=[
        {
            "role": "user",
            "content": user_input,
            }
        ]
    steps=[]
    trace_path = (create_trace_path() if trace else None)

    def finish(result: dict) -> dict:
        if trace_path is not None:
            result["trace_file"] = str(trace_path)
        append_trace_event(
            trace_path,
            {
                "event": "run_finished",
                "result": result,
            },
        )
        return result

    for step_number in range(1, max_steps + 1):
        #Logging convo given to agent_step
        append_trace_event(
            trace_path,
            {
                "event": "agent_step_input",
                "step_number": step_number,
                "messages": messages,
            },
        )
        step = agent_step(
            llm_config=llm_config,
            messages=messages,
            system_prompt=system_prompt,
            max_tokens=max_tokens,
            temperature=temperature,
        )
        append_trace_event(
            trace_path,
            {
                "event": "agent_step_output",
                "step_number": step_number,
                "step": step,
            },
        )
        if step is None:
            return finish({
                "answer": "I don't know.",
                "documents_used": [],
                "steps": steps,
                "error": (
                    "Agent failed to return a valid action."
                ),
            })
        action = step.get("action")
        if action == "final_answer":
            if not get_successful_tool_messages(messages):
                return finish({
                    "answer": "I don't know.",
                    "documents_used": [],
                    "steps": steps,
                    "error": (
                        "Agent attempted to answer before "
                        "a tool completed successfully."
                    ),
                })
            steps.append(step)
            evidence_ids = step.get("evidence_ids", [])
            documents_used = get_documents_used(
                messages=messages,
                evidence_ids=evidence_ids,
            )
            return finish({
                "answer": step.get(
                    "answer",
                    "I don't know.",
                ),
                "documents_used": documents_used,
                "steps": steps,
            })
        if action == "tool_use":
            steps.append(step)
            tool_result = execute_tool_call(
                tool_call=step,
                vectorstore=vectorstore,
            )
            assistant_message = {
                "role": "assistant",
                "content": json.dumps(
                    step,
                    ensure_ascii=False,
                ),
            }
            tool_message = {
                "role": "tool",
                "name": step.get("tool"),
                "success": tool_result.get(
                    "success",
                    False,
                ),
                "content": json.dumps(
                    tool_result,
                    ensure_ascii=False,
                ),
            }
            messages.extend([assistant_message, tool_message])
            append_trace_event(
                trace_path,
                {
                    "event": "tool_result",
                    "step_number": step_number,
                    "tool": step.get("tool"),
                    "result": tool_result,
                },
            )
            if not tool_result.get("success", False):
                return finish({
                    "answer": "I don't know.",
                    "documents_used": [],
                    "steps": steps,
                    "error": (
                        f"Tool '{step.get('tool')}' failed: "
                        f"{tool_result.get('error')}"
                    ),
                })
            continue
        steps.append(step)
        return finish({
            "answer": "I don't know.",
            "documents_used": [],
            "steps": steps,
            "error": f"Unknown action: {action}",
        })
    return finish({
        "answer": "I don't know.",
        "documents_used": [],
        "steps": steps,
        "error": (
            "Reached max_steps before producing "
            "a final answer."
        ),
    })

if __name__ == "__main__":
    main()
